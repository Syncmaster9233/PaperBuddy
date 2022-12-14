from tkinter import *
from openpyxl import Workbook, load_workbook
import csv
import datetime
import os
import shutil

class Root(Tk):
	def __init__(self):
		super(Root,self).__init__()

		self.title("Paper Buddy")
		self.minsize(500,350)
		self.iconbitmap('Resources/Icon.ico')
		self['background'] = '#444444'

root = Root()

Name = StringVar()
Job = StringVar()
Model = StringVar()
Part = StringVar()
errMess = StringVar()

#getting saved settings from last job
settingsFile = open('Resources/Settings.txt', 'r')
setraw = settingsFile.readlines()
settingsFile.close()

Name = setraw[0].replace('\n', '')
Job = setraw[1].replace('\n', '')
Model = setraw[2].replace('\n', '')
Part = setraw[3].replace('\n', '')


def MachineShopGen():
	try:
		errMess.set("Generating")
		#opening master template
		wb = load_workbook('Resources/MST.xlsx')

		#getting raw data for parts list going to machine shop
		partcount = 0
		partlist = []
		with open('C:/Users/' + os.getlogin() + '/Downloads/SelectionList.csv') as csv_file:
			#parsing through rows in csv export
			for row in csv_file:

				#skipping first row
				if partcount > 0:
					#parsing row into strings
					rowArr = row.split(",")
					
					#grabbing only info needed part numbers, name and quantity into impArr
					impArr = []
					blockTrack = 0
					for block in rowArr:
						if blockTrack > 0 and blockTrack < 4:
							impArr.append(block)
						blockTrack+=1

					#appending impArr into partlist
					partlist.append(impArr)
				#tracking rows with parts
				partcount +=1	

		#consolidating parts
		#getting part numbers for unique parts
		uniqpartlist = []
		for part in partlist:
			unique = True
			for upart in uniqpartlist:
				if part[0] == upart[0]:
					unique = False
			if unique:
				uniqpartlist.append(part)

		#summing all unique parts
		#probably can be done in previous loop but too lazy
		for part in uniqpartlist:
			sum = 0
			for pt in partlist:
				if part[0] == pt[0]:
					sum += int(pt[2])
			part[2] = sum

		#finding required number of sheets
		uniqPartCount = len(uniqpartlist)
		sheetNum = int(uniqPartCount/4) 
		if uniqPartCount%4 > 0:
			sheetNum += 1

		#getting inputs
		name = entryN.get()
		jobNum = entryJ.get()
		model = entryM.get()
		part = entryP.get()

		with open('Resources/Settings.txt','w') as s:
			s.write(name + '\n')
			s.write(jobNum + '\n')
			s.write(model + '\n')
			s.write(part + '\n')

		#getting date
		dateTemp = datetime.datetime.now()
		date = str(dateTemp.month) + '-' + str(dateTemp.day) + '-' + str(dateTemp.year)

		Msheet = wb.active

		#input settings and date into master template
		Msheet['A2'] = str(date)
		Msheet['A7'] = str(jobNum)
		Msheet['J7'] = str(model)
		Msheet['M7'] = str(part)
		Msheet['K11'] = str(name)
		Msheet['K18'] = str(name)
		Msheet['K25'] = str(name)
		Msheet['K32'] = str(name)

		#copying base sheets for req sheet count
		for x in range(1,sheetNum,1):
			wb.copy_worksheet(Msheet)

		#cleaning up sheet numbers and page nums and exporting data to spreed sheet
		sheetct = 1
		for sheet in wb.worksheets:

			#cleaning up names
			sheet.title = 'Sheet'+str(sheetct)
			#page nums input
			sheet['L2'] = 'Page ' + str(sheetct) + ' of ' + str(sheetNum)

			begin = (sheetct-1)*4

			try:
				sheet['A14'] = uniqpartlist[begin][0]
				sheet['A11'] = uniqpartlist[begin][1] + ' (' + str(uniqpartlist[begin][2]) + ')'
			except:
				print('end of uniqpartlist')
			try:
				sheet['A21'] = uniqpartlist[begin+1][0]
				sheet['A18'] = uniqpartlist[begin+1][1] + ' (' + str(uniqpartlist[begin+1][2]) + ')'
			except:
				print('end of uniqpartlist')
			try:
				sheet['A28'] = uniqpartlist[begin+2][0]
				sheet['A25'] = uniqpartlist[begin+2][1] + ' (' + str(uniqpartlist[begin+2][2]) + ')'
			except:
				print('end of uniqpartlist')
			try:
				sheet['A35'] = uniqpartlist[begin+3][0]
				sheet['A32'] = uniqpartlist[begin+3][1] + ' (' + str(uniqpartlist[begin+3][2]) + ')'
			except:
				print('end of uniqpartlist')

			sheetct += 1

		wb.save('C:/Users/' + os.getlogin() + '/Desktop/' + jobNum + '.xlsx')
		try:
			shutil.move('C:/Users/' + os.getlogin() + '/Downloads/SelectionList.csv', 'Resources/')
		except:
			os.remove('Resources/SelectionList.csv')
			shutil.move('C:/Users/' + os.getlogin() + '/Downloads/SelectionList.csv', 'Resources/')
		errMess.set("Generated")
		wb.close()
	except FileNotFoundError:
		errMess.set("Error: put the exported list in downloads folder")

def MachineShopReGen():
	try:
		errMess.set("ReGenerating")
		#opening master template
		wb = load_workbook('Resources/MST.xlsx')

		#getting raw data for parts list going to machine shop
		partcount = 0
		partlist = []
		with open('Resources/SelectionList.csv') as csv_file:
			#parsing through rows in csv export
			for row in csv_file:

				#skipping first row
				if partcount > 0:
					#parsing row into strings
					rowArr = row.split(",")
					
					#grabbing only info needed part numbers, name and quantity into impArr
					impArr = []
					blockTrack = 0
					for block in rowArr:
						if blockTrack > 0 and blockTrack < 4:
							impArr.append(block)
						blockTrack+=1

					#appending impArr into partlist
					partlist.append(impArr)
				#tracking rows with parts
				partcount +=1	

		#consolidating parts
		#getting part numbers for unique parts
		uniqpartlist = []
		for part in partlist:
			unique = True
			for upart in uniqpartlist:
				if part[0] == upart[0]:
					unique = False
			if unique:
				uniqpartlist.append(part)

		#summing all unique parts
		#probably can be done in previous loop but too lazy
		for part in uniqpartlist:
			sum = 0
			for pt in partlist:
				if part[0] == pt[0]:
					sum += int(pt[2])
			part[2] = sum

		#finding required number of sheets
		uniqPartCount = len(uniqpartlist)
		sheetNum = int(uniqPartCount/4) 
		if uniqPartCount%4 > 0:
			sheetNum += 1

		#getting inputs
		name = entryN.get()
		jobNum = entryJ.get()
		model = entryM.get()
		part = entryP.get()

		with open('Resources/Settings.txt','w') as s:
			s.write(name + '\n')
			s.write(jobNum + '\n')
			s.write(model + '\n')
			s.write(part + '\n')

		#getting date
		dateTemp = datetime.datetime.now()
		date = str(dateTemp.month) + '-' + str(dateTemp.day) + '-' + str(dateTemp.year)

		Msheet = wb.active

		#input settings and date into master template
		Msheet['A2'] = str(date)
		Msheet['A7'] = str(jobNum)
		Msheet['J7'] = str(model)
		Msheet['M7'] = str(part)
		Msheet['K11'] = str(name)
		Msheet['K18'] = str(name)
		Msheet['K25'] = str(name)
		Msheet['K32'] = str(name)

		#copying base sheets for req sheet count
		for x in range(1,sheetNum,1):
			wb.copy_worksheet(Msheet)

		#cleaning up sheet numbers and page nums and exporting data to spreed sheet
		sheetct = 1
		for sheet in wb.worksheets:

			#cleaning up names
			sheet.title = 'Sheet'+str(sheetct)
			#page nums input
			sheet['L2'] = 'Page ' + str(sheetct) + ' of ' + str(sheetNum)

			begin = (sheetct-1)*4

			try:
				sheet['A14'] = uniqpartlist[begin][0]
				sheet['A11'] = uniqpartlist[begin][1] + ' (' + str(uniqpartlist[begin][2]) + ')'
			except:
				print('end of uniqpartlist')
			try:
				sheet['A21'] = uniqpartlist[begin+1][0]
				sheet['A18'] = uniqpartlist[begin+1][1] + ' (' + str(uniqpartlist[begin+1][2]) + ')'
			except:
				print('end of uniqpartlist')
			try:
				sheet['A28'] = uniqpartlist[begin+2][0]
				sheet['A25'] = uniqpartlist[begin+2][1] + ' (' + str(uniqpartlist[begin+2][2]) + ')'
			except:
				print('end of uniqpartlist')
			try:
				sheet['A35'] = uniqpartlist[begin+3][0]
				sheet['A32'] = uniqpartlist[begin+3][1] + ' (' + str(uniqpartlist[begin+3][2]) + ')'
			except:
				print('end of uniqpartlist')

			sheetct += 1

		wb.save('C:/Users/' + os.getlogin() + '/Desktop/' + jobNum + '.xlsx')
		errMess.set("ReGenerated")
		wb.close()
	except FileNotFoundError:
		errMess.set("Error: no prior history of generation")

#creating main frame
mainFrm = Frame(root,bg = '#444444')

#creating entry frame and widgets
entryFrm = Frame(mainFrm,bg = '#444444')

labelN = Label(entryFrm, text = "Name:",bg = '#444444',fg = 'white')
labelJ = Label(entryFrm, text = "Job#:",bg = '#444444',fg = 'white')
labelM = Label(entryFrm, text = "Model:",bg = '#444444',fg = 'white')
labelP = Label(entryFrm, text = "Part:",bg = '#444444',fg = 'white')

entryN = Entry(entryFrm, width = 50,bg = '#444444',fg = 'white')
entryJ = Entry(entryFrm, width = 50,bg = '#444444',fg = 'white')
entryM = Entry(entryFrm, width = 50,bg = '#444444',fg = 'white')
entryP = Entry(entryFrm, width = 50,bg = '#444444',fg = 'white')

#structuring entry frame
labelN.grid(row = 0,column = 0, pady = 5)
entryN.grid(row = 0,column = 1)
labelJ.grid(row = 1,column = 0, pady = 5)
entryJ.grid(row = 1,column = 1)
labelM.grid(row = 2,column = 0, pady = 5)
entryM.grid(row = 2,column = 1)
labelP.grid(row = 3,column = 0, pady = 5)
entryP.grid(row = 3,column = 1)

#inputing settings history into input fields
entryN.insert(END,Name)
entryJ.insert(END,Job)
entryM.insert(END,Model)
entryP.insert(END,Part)

#creating button frame and buttons
buttonfrm = Frame(mainFrm, bg = '#444444')
buttonGen = Button(buttonfrm, text="Generate Form", command = MachineShopGen,bg = '#444444',fg = 'white')
buttonReGen = Button(buttonfrm, text = "Regenerate Form", command = MachineShopReGen,bg = '#444444',fg = 'white')

#structuring buttons
buttonGen.grid(row = 4,column = 0, pady = 5, padx = 30)
buttonReGen.grid(row = 4,column = 1, pady = 5, padx = 30)

errlabel = Label(mainFrm, textvariable = errMess,bg = '#444444',fg = 'orange')

insLabel = Label(mainFrm, text = """	Note: Program will move exported list when generating to prevent over crowding
	Program will generate the new excel file on the desktop. The user still needs to
	input work requested. Program will remember settings and previous exported list 
	on open and close. Regenerate to change settings without exporting list again.
	Regeneration will erase work requested and if job number is changed create a new
	file entirely.
	""",bg = '#444444',fg = 'white')

entryFrm.pack()
buttonfrm.pack()
errlabel.pack()
insLabel.pack()

mainFrm.pack()

root.mainloop()

